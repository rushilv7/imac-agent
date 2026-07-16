from __future__ import annotations

import csv
import hashlib
import io
import json
import mimetypes
import os
import sys
from collections import Counter
from pathlib import Path
from typing import Any

THIS_DIR = Path(__file__).resolve().parent
if str(THIS_DIR) not in sys.path:
    sys.path.insert(0, str(THIS_DIR))

from config import INCOMING_DIR, KNOWLEDGE_ROOT, ensure_directories
from registry import get_duplicate_by_sha256, register

TEXT_LIMIT = 40000


class IngestError(RuntimeError):
    pass


def _validated_incoming_path(path: str | Path) -> Path:
    ensure_directories()
    resolved = Path(path).expanduser().resolve()
    incoming_root = INCOMING_DIR.resolve()
    try:
        resolved.relative_to(incoming_root)
    except ValueError as exc:
        raise IngestError("Path must resolve inside ~/knowledge/incoming") from exc

    if not resolved.is_file():
        raise IngestError("Path is not a file")

    return resolved


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _guess_mime(path: Path) -> str:
    mime, _ = mimetypes.guess_type(str(path))
    return mime or "application/octet-stream"


def _safe_text(value: str) -> str:
    if not value:
        return ""
    return value[:TEXT_LIMIT]


def _deterministic_category(extension: str, mime_type: str) -> str:
    ext = extension.lower().lstrip(".")
    if ext in {"csv", "tsv"}:
        return "data"
    if ext in {"xlsx", "xlsm", "xls"}:
        return "data"
    if ext in {"png", "jpg", "jpeg", "gif", "webp", "tiff", "bmp"}:
        return "images"
    if ext in {"pdf", "txt", "md", "json", "yaml", "yml", "log"}:
        return "documents"
    if mime_type.startswith("text/"):
        return "documents"
    return "other"


def _extract_pdf(path: Path) -> tuple[dict[str, Any], str]:
    metadata: dict[str, Any] = {"kind": "pdf"}
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as exc:  # pragma: no cover
        metadata["error"] = f"pypdf unavailable: {type(exc).__name__}"
        return metadata, ""

    try:
        reader = PdfReader(str(path))
        page_count = len(reader.pages)
        metadata["page_count"] = page_count

        text_parts: list[str] = []
        remaining = TEXT_LIMIT
        for page in reader.pages[: max(0, min(page_count, 9999))]:
            if remaining <= 0:
                break
            extracted = page.extract_text() or ""
            if extracted:
                snippet = extracted[:remaining]
                text_parts.append(snippet)
                remaining -= len(snippet)

        return metadata, "\n".join(text_parts)
    except Exception as exc:
        metadata["error"] = f"PDF extraction failed: {type(exc).__name__}"
        return metadata, ""


def _detect_csv_dialect(sample: str) -> csv.Dialect:
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample, delimiters=",\t;|")
    except Exception:
        dialect = csv.get_dialect("excel")
    return dialect  # type: ignore[return-value]


def _extract_csv_like(path: Path) -> tuple[dict[str, Any], str]:
    text = path.read_text("utf-8", errors="replace")
    sample = text[: min(len(text), 20000)]
    dialect = _detect_csv_dialect(sample)

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows: list[list[str]] = []
    for index, row in enumerate(reader):
        if index >= 5000:
            break
        rows.append(row)

    metadata: dict[str, Any] = {
        "kind": "csv",
        "delimiter": getattr(dialect, "delimiter", ","),
        "row_count": len(rows),
    }

    headers: list[str] = []
    if rows:
        headers = [cell.strip() for cell in rows[0]]
        metadata["headers"] = headers

    # Null counts by column.
    null_counts: dict[str, int] = {}
    if headers:
        for col_index, name in enumerate(headers):
            count = 0
            for row in rows[1:]:
                value = row[col_index] if col_index < len(row) else ""
                if value == "" or value is None:
                    count += 1
            null_counts[name or f"col_{col_index+1}"] = count
        metadata["null_counts"] = null_counts

    # Exact duplicate count across full rows (excluding header).
    duplicates = 0
    if len(rows) > 2:
        body_rows = [tuple(r) for r in rows[1:]]
        counts = Counter(body_rows)
        duplicates = sum((c - 1) for c in counts.values() if c > 1)
    metadata["exact_duplicate_rows"] = duplicates

    sample_rows = rows[:31]  # header + 30
    sample_out = io.StringIO()
    writer = csv.writer(sample_out, dialect=dialect)
    writer.writerows(sample_rows)
    return metadata, sample_out.getvalue()[:TEXT_LIMIT]


def _extract_xlsx(path: Path) -> tuple[dict[str, Any], str]:
    metadata: dict[str, Any] = {"kind": "xlsx"}
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        metadata["error"] = f"openpyxl unavailable: {type(exc).__name__}"
        return metadata, ""

    try:
        workbook = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        metadata["error"] = f"Workbook load failed: {type(exc).__name__}"
        return metadata, ""

    sheet_names = workbook.sheetnames[:10]
    metadata["sheets"] = []

    out = io.StringIO()
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]

        max_rows = 5000
        max_cols = 200
        rows: list[list[Any]] = []
        for r_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_index >= max_rows:
                break
            rows.append(list(row[:max_cols]))

        headers: list[str] = []
        if rows:
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

        null_counts: dict[str, int] = {}
        if headers:
            for c_index, name in enumerate(headers):
                count = 0
                for row in rows[1:]:
                    value = row[c_index] if c_index < len(row) else None
                    if value is None or value == "":
                        count += 1
                null_counts[name or f"col_{c_index+1}"] = count

        duplicates = 0
        if len(rows) > 2:
            body_rows = [tuple(r) for r in rows[1:]]
            counts = Counter(body_rows)
            duplicates = sum((c - 1) for c in counts.values() if c > 1)

        metadata["sheets"].append(
            {
                "name": sheet_name,
                "row_count": len(rows),
                "col_count": max((len(r) for r in rows), default=0),
                "headers": headers,
                "null_counts": null_counts,
                "exact_duplicate_rows": duplicates,
            }
        )

        out.write(f"===== SHEET: {sheet_name} =====\n")
        for row in rows[:31]:
            out.write("\t".join("" if cell is None else str(cell) for cell in row))
            out.write("\n")
        out.write("\n")

    return metadata, out.getvalue()[:TEXT_LIMIT]


def _extract_image(path: Path) -> tuple[dict[str, Any], str]:
    metadata: dict[str, Any] = {"kind": "image"}
    try:
        from PIL import Image  # type: ignore
    except Exception as exc:  # pragma: no cover
        metadata["error"] = f"Pillow unavailable: {type(exc).__name__}"
        return metadata, ""

    try:
        with Image.open(path) as image:
            metadata["format"] = image.format
            metadata["mode"] = image.mode
            metadata["width"] = image.width
            metadata["height"] = image.height

            exif_safe: dict[str, Any] = {}
            try:
                exif = image.getexif()
                if exif:
                    # Only keep simple scalars.
                    for key, value in dict(exif).items():
                        if isinstance(value, (str, int, float)):
                            exif_safe[str(key)] = value
            except Exception:
                pass
            if exif_safe:
                metadata["exif"] = exif_safe

        return metadata, ""
    except Exception as exc:
        metadata["error"] = f"Image read failed: {type(exc).__name__}"
        return metadata, ""


def ingest(path: str | Path) -> dict[str, Any]:
    file_path = _validated_incoming_path(path)

    stat = file_path.stat()
    size_bytes = int(stat.st_size)
    extension = file_path.suffix.lower().lstrip(".")
    mime_type = _guess_mime(file_path)
    digest = _sha256(file_path)

    duplicate = get_duplicate_by_sha256(digest)
    if duplicate:
        return {
            "ok": True,
            "duplicate": True,
            "knowledge_item_id": int(duplicate["id"]),
            "sha256": digest,
            "mime_type": mime_type,
            "size_bytes": size_bytes,
            "extension": extension,
            "category": duplicate.get("category"),
        }

    category = _deterministic_category(extension, mime_type)

    metadata: dict[str, Any] = {
        "mime_type": mime_type,
        "size_bytes": size_bytes,
        "extension": extension,
        "category": category,
    }

    extracted_text = ""
    if extension == "pdf" or mime_type == "application/pdf":
        extra, extracted_text = _extract_pdf(file_path)
        metadata.update(extra)
        extracted_text = _safe_text(extracted_text)
    elif extension in {"csv", "tsv"} or mime_type in {"text/csv", "application/csv"}:
        extra, sample = _extract_csv_like(file_path)
        metadata.update(extra)
        metadata["sample"] = sample
    elif extension in {"xlsx", "xlsm", "xls"}:
        extra, sample = _extract_xlsx(file_path)
        metadata.update(extra)
        metadata["sample"] = sample
    elif extension in {"txt", "md", "json", "yaml", "yml", "log", "tsv"} or mime_type.startswith("text/"):
        extracted_text = _safe_text(file_path.read_text("utf-8", errors="replace"))
        metadata["text_sample"] = extracted_text
    elif category == "images":
        extra, _ = _extract_image(file_path)
        metadata.update(extra)
    else:
        metadata["kind"] = "unknown"

    if extracted_text and "text_sample" not in metadata:
        metadata["text_sample"] = extracted_text

    # Phase 2: store a bounded extracted_text for full-text search and enrichment.
    extracted_for_index = ""
    if extracted_text:
        extracted_for_index = extracted_text
    else:
        sample = metadata.get("text_sample") or metadata.get("sample")
        if isinstance(sample, str):
            extracted_for_index = sample[:TEXT_LIMIT]

    item = register(
        stored_path=file_path,
        sha256=digest,
        original_name=file_path.name,
        mime_type=mime_type,
        size_bytes=size_bytes,
        extension=extension,
        category=category,
        status="ingested",
        summary=None,
        keywords=[],
        metadata=metadata,
        extracted_text=extracted_for_index,
        suggested_path=None,
    )

    return {
        "ok": True,
        "duplicate": False,
        "knowledge_item_id": int(item["id"]),
        "sha256": digest,
        "mime_type": mime_type,
        "size_bytes": size_bytes,
        "extension": extension,
        "category": category,
    }
