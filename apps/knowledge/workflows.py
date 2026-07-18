from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from config import ARTIFACTS_DIR, KNOWLEDGE_ROOT, ensure_directories
from artifacts import create_artifact
from registry import get as get_item


class WorkflowError(RuntimeError):
    pass


ALLOWED_OPERATIONS = {
    "trim_strings",
    "normalize_booleans",
    "normalize_dates",
    "remove_exact_duplicates",
    "standardize_column_names",
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_operations(spec: str) -> list[str]:
    raw = (spec or "").strip()
    if not raw:
        return []
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    seen: set[str] = set()
    ops: list[str] = []
    for part in parts:
        if part in seen:
            continue
        seen.add(part)
        ops.append(part)
    return ops


def validate_operations(operations: list[str]) -> list[str]:
    invalid = [op for op in operations if op not in ALLOWED_OPERATIONS]
    if invalid:
        raise WorkflowError(
            "Unsupported workflow operations: " + ", ".join(sorted(set(invalid)))
        )
    return operations


def _validated_under_root(path: Path, root: Path) -> Path:
    resolved = path.expanduser().resolve()
    root_resolved = root.expanduser().resolve()
    try:
        resolved.relative_to(root_resolved)
    except ValueError as exc:
        raise WorkflowError("Path is outside allowlisted root") from exc
    return resolved


def validated_knowledge_item_path(item: dict[str, Any]) -> Path:
    path = Path(str(item.get("stored_path") or "")).expanduser().resolve()
    root = KNOWLEDGE_ROOT.resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise WorkflowError("Knowledge item path is outside knowledge root") from exc
    if not path.is_file():
        raise WorkflowError("Knowledge item source file is missing")
    return path


def _unique_destination(dest_dir: Path, filename: str) -> Path:
    dest_dir.mkdir(parents=True, exist_ok=True)
    base = Path(filename).stem
    suffix = Path(filename).suffix

    candidate = dest_dir / f"{base}{suffix}"
    if not candidate.exists():
        return candidate

    for index in range(2, 1000):
        candidate = dest_dir / f"{base}_{index}{suffix}"
        if not candidate.exists():
            return candidate

    raise WorkflowError("Could not find a unique artifact filename")


def _read_csv_all_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    text = path.read_text("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader]
    if not rows:
        return [], []
    header = [cell for cell in rows[0]]
    body = [list(row) for row in rows[1:]]
    return header, body


def _read_xlsx_all_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise WorkflowError(f"openpyxl is required for Excel workflows: {type(exc).__name__}") from exc

    try:
        workbook = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise WorkflowError(f"Workbook load failed: {type(exc).__name__}") from exc

    if not workbook.sheetnames:
        return [], []
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[list[str]] = []
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx >= 200000:
            break
        rows.append(["" if cell is None else str(cell) for cell in row])

    if not rows:
        return [], []
    header = [cell for cell in rows[0]]
    body = [list(r) for r in rows[1:]]
    return header, body


def _write_xlsx(path: Path, header: list[str], body: list[list[str]]) -> None:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise WorkflowError(f"openpyxl is required for Excel workflows: {type(exc).__name__}") from exc

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:  # pragma: no cover
        raise WorkflowError("Workbook did not provide an active sheet")
    sheet.title = "cleaned"

    sheet.append([cell for cell in header])
    for row in body:
        sheet.append([cell for cell in row])

    try:
        workbook.save(str(path))
    except Exception as exc:
        raise WorkflowError(f"Workbook save failed: {type(exc).__name__}") from exc


def _write_csv(path: Path, header: list[str], body: list[list[str]]) -> None:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(body)
    data = output.getvalue().encode("utf-8")
    path.write_bytes(data)


def _apply_trim_strings(header: list[str], body: list[list[str]]) -> None:
    for r_idx, row in enumerate(body):
        body[r_idx] = [cell.strip() if isinstance(cell, str) else cell for cell in row]


def _apply_standardize_column_names(header: list[str]) -> list[str]:
    out: list[str] = []
    seen: dict[str, int] = {}
    for value in header:
        raw = (value or "").strip().lower()
        raw = re.sub(r"[^a-z0-9]+", "_", raw).strip("_")
        name = raw or "column"
        count = seen.get(name, 0) + 1
        seen[name] = count
        if count > 1:
            name = f"{name}_{count}"
        out.append(name)
    return out


_TRUE = {"true", "t", "yes", "y", "1"}
_FALSE = {"false", "f", "no", "n", "0"}


def _apply_normalize_booleans(body: list[list[str]]) -> None:
    for r_idx, row in enumerate(body):
        new_row: list[str] = []
        for cell in row:
            if isinstance(cell, str):
                lower = cell.strip().lower()
                if lower in _TRUE:
                    new_row.append("true")
                    continue
                if lower in _FALSE:
                    new_row.append("false")
                    continue
            new_row.append(cell)
        body[r_idx] = new_row


_DATE_PATTERNS = (
    (re.compile(r"^(\d{4})-(\d{2})-(\d{2})$"), "%Y-%m-%d"),
    (re.compile(r"^(\d{4})/(\d{2})/(\d{2})$"), "%Y/%m/%d"),
    (re.compile(r"^(\d{2})/(\d{2})/(\d{4})$"), "%m/%d/%Y"),
)


def _normalize_date(cell: str) -> str:
    raw = cell.strip()
    for pattern, fmt in _DATE_PATTERNS:
        if not pattern.match(raw):
            continue
        try:
            parsed = datetime.strptime(raw, fmt)
        except Exception:
            continue
        return parsed.strftime("%Y-%m-%d")
    return cell


def _apply_normalize_dates(body: list[list[str]]) -> None:
    for r_idx, row in enumerate(body):
        new_row: list[str] = []
        for cell in row:
            if isinstance(cell, str):
                new_row.append(_normalize_date(cell))
            else:
                new_row.append(cell)
        body[r_idx] = new_row


def _apply_remove_exact_duplicates(body: list[list[str]]) -> int:
    seen: set[tuple[str, ...]] = set()
    out: list[list[str]] = []
    removed = 0
    for row in body:
        key = tuple(str(cell) for cell in row)
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        out.append(row)
    body[:] = out
    return removed


@dataclass
class WorkflowReport:
    rows_before: int
    rows_after: int
    duplicate_rows_removed: int
    columns_before: list[str]
    columns_after: list[str]
    operations: list[str]
    artifact_id: int
    artifact_path: str


def run_workflow(
    *,
    knowledge_item_id: int,
    operations: list[str],
) -> WorkflowReport:
    ensure_directories()

    operations = validate_operations(list(operations))

    item = get_item(int(knowledge_item_id))
    if not item:
        raise WorkflowError("Knowledge item not found")

    source_path = validated_knowledge_item_path(item)
    ext = source_path.suffix.lower()
    if ext in {".csv", ".tsv"}:
        header, body = _read_csv_all_rows(source_path)
        artifact_mime = "text/csv"
        base_name = f"cleaned_{source_path.name}"
        write_output = lambda dest, h, b: _write_csv(dest, h, b)
    elif ext in {".xlsx", ".xlsm"}:
        header, body = _read_xlsx_all_rows(source_path)
        artifact_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        base_name = f"cleaned_{source_path.stem}.xlsx"
        write_output = lambda dest, h, b: _write_xlsx(dest, h, b)
    else:
        raise WorkflowError("Only CSV/TSV/XLSX workflows are supported in this implementation")
    columns_before = list(header)
    rows_before = len(body)

    if "trim_strings" in operations:
        _apply_trim_strings(header, body)

    if "normalize_booleans" in operations:
        _apply_normalize_booleans(body)

    if "normalize_dates" in operations:
        _apply_normalize_dates(body)

    duplicate_rows_removed = 0
    if "remove_exact_duplicates" in operations:
        duplicate_rows_removed = _apply_remove_exact_duplicates(body)

    if "standardize_column_names" in operations:
        header = _apply_standardize_column_names(header)

    rows_after = len(body)
    columns_after = list(header)

    # Write artifact.
    ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True, mode=0o700)
    artifact_path = _unique_destination(ARTIFACTS_DIR, base_name)
    artifact_path = _validated_under_root(artifact_path, ARTIFACTS_DIR)

    write_output(artifact_path, header, body)

    artifact = create_artifact(
        source_item_ids=[int(knowledge_item_id)],
        filename=artifact_path.name,
        stored_path=artifact_path,
        mime_type=artifact_mime,
        description=(
            "Cleaned dataset artifact generated from knowledge item "
            f"#{knowledge_item_id} using operations: {', '.join(operations)}"
        ),
    )

    return WorkflowReport(
        rows_before=rows_before,
        rows_after=rows_after,
        duplicate_rows_removed=duplicate_rows_removed,
        columns_before=columns_before,
        columns_after=columns_after,
        operations=operations,
        artifact_id=int(artifact["id"]),
        artifact_path=str(artifact_path),
    )
